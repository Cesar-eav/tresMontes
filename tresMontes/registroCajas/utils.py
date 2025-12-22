"""
Utilidades para el sistema de registro de cajas
"""
import openpyxl
import csv
import io
from .models import Beneficiario, Planta


def procesar_excel_nomina(archivo, campana, planta):
    """
    Procesa un archivo Excel o CSV con la nómina de beneficiarios

    Formato esperado:
    - Columna 1: Nombre completo
    - Columna 2: RUT (formato: 12.345.678-9)
    - Columna 3: Tipo de contrato (planta/contratista)
    - Columna 4 (opcional): Tipo de caja (estandar/especial/premium)

    Returns:
        int: Número de beneficiarios creados
    """
    try:
        # Detectar si es CSV o Excel
        nombre_archivo = archivo.name.lower()

        if nombre_archivo.endswith('.csv'):
            return _procesar_csv_nomina(archivo, campana, planta)
        else:
            return _procesar_excel_nomina(archivo, campana, planta)

    except Exception as e:
        raise Exception(f"Error al procesar archivo: {str(e)}")


def _procesar_csv_nomina(archivo, campana, planta):
    """
    Procesa un archivo CSV con formato flexible:

    Formato simplificado (5 columnas):
    RUT | NOMBRE | TIPO_CONTRATO | TIPO_CAJA | PLANTA_ID

    Formato extendido (9 columnas - legacy):
    RUT | EMPLEADO | NOMBRES | APELLIDOS | CARGO | TIPO DE CONTRATO | PERIODO | SEDE | ESTADO
    """
    beneficiarios_creados = 0
    filas_procesadas = 0
    errores = []

    try:
        # Leer el archivo CSV
        archivo_texto = archivo.read().decode('utf-8-sig')  # utf-8-sig maneja BOM
        print(f"DEBUG: Archivo leído, {len(archivo_texto)} caracteres")

        # Validar que el archivo no esté vacío
        if not archivo_texto or archivo_texto.strip() == '':
            raise Exception("El archivo está vacío. Por favor suba un archivo con datos.")

        # Detectar delimitador (tabulador o coma)
        if '\t' in archivo_texto:
            lector_csv = csv.reader(io.StringIO(archivo_texto), delimiter='\t')
            print(f"DEBUG: Usando delimitador TABULADOR")
        else:
            lector_csv = csv.reader(io.StringIO(archivo_texto))
            print(f"DEBUG: Usando delimitador COMA")

        # Leer encabezado
        header = next(lector_csv, None)
        print(f"DEBUG: Encabezado CSV: {header}")
        print(f"DEBUG: Número de columnas en encabezado: {len(header) if header else 0}")

        # Validar que exista encabezado
        if not header:
            raise Exception("El archivo no tiene encabezado. Asegúrese de que la primera fila contenga los nombres de las columnas.")

        # Validar número de columnas (debe tener al menos 4)
        num_columnas = len(header)
        if num_columnas < 4:
            raise Exception(f"El archivo tiene {num_columnas} columnas pero se requieren al menos 4 columnas: RUT, NOMBRE, TIPO_CONTRATO, TIPO_CAJA/PLANTA_ID")

        # Detectar formato del CSV basado en número de columnas
        formato_simplificado = num_columnas >= 4 and num_columnas <= 5
        print(f"DEBUG: Formato detectado: {'SIMPLIFICADO (4-5 columnas)' if formato_simplificado else 'EXTENDIDO (9+ columnas)'}")

        # Detectar índices de columnas importantes
        header_lower = [h.strip().lower() if h else '' for h in header]

        # Detectar índice de planta/sede
        planta_idx = None
        for i, h in enumerate(header_lower):
            if any(k in h for k in ['planta', 'planta_id', 'sede', 'site', 'sucursal', 'centro']):
                planta_idx = i
                break

        # Detectar índice de tipo de contrato
        tipo_contrato_idx = None
        for i, h in enumerate(header_lower):
            if any(k in h for k in ['tipo', 'contrato', 'tipo_contrato']):
                tipo_contrato_idx = i
                break

        print(f"DEBUG: Índice de columna planta detectado: {planta_idx}")
        print(f"DEBUG: Índice de columna tipo_contrato detectado: {tipo_contrato_idx}")

        for idx, row in enumerate(lector_csv, start=2):
            filas_procesadas += 1
            print(f"DEBUG: Fila {idx}: {row}")
            print(f"DEBUG: Longitud de fila: {len(row)}")

            # Saltar filas completamente vacías
            if not row or all(not str(cell).strip() for cell in row):
                print(f"DEBUG: Fila {idx} está vacía, saltando...")
                continue

            # Validar que tenga al menos las columnas mínimas
            min_cols = 4
            if len(row) < min_cols:
                error = f"Fila {idx}: Tiene {len(row)} columnas pero se requieren al menos {min_cols}"
                print(f"DEBUG ERROR: {error}")
                errores.append(error)
                continue

            # Procesar según el formato detectado
            if formato_simplificado:
                # Formato simplificado: RUT | NOMBRE | TIPO_CONTRATO | TIPO_CAJA | PLANTA_ID
                # 0: RUT
                # 1: NOMBRE
                # 2: TIPO_CONTRATO
                # 3: TIPO_CAJA (opcional, por defecto 'estandar')
                # 4: PLANTA_ID (opcional, usa planta por defecto si no se especifica)

                rut = str(row[0]).strip() if row[0] else ''
                nombre_completo = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                tipo_contrato_raw = str(row[2]).strip().lower() if len(row) > 2 and row[2] else 'indefinido'
                tipo_caja_raw = str(row[3]).strip().lower() if len(row) > 3 and row[3] else 'estandar'
                planta_id_raw = str(row[4]).strip() if len(row) > 4 and row[4] else ''

                print(f"DEBUG: [FORMATO SIMPLIFICADO] RUT={rut}, NOMBRE={nombre_completo}, TIPO_CONTRATO={tipo_contrato_raw}, TIPO_CAJA={tipo_caja_raw}, PLANTA_ID={planta_id_raw}")
            else:
                # Formato extendido: RUT | EMPLEADO | NOMBRES | APELLIDOS | CARGO | TIPO DE CONTRATO | PERIODO | SEDE | ESTADO
                # 0: RUT
                # 1: EMPLEADO
                # 2: NOMBRES
                # 3: APELLIDOS
                # 4: CARGO (ignorado)
                # 5: TIPO DE CONTRATO
                # 6: PERIODO (ignorado)
                # 7: SEDE
                # 8: ESTADO (ignorado)

                rut = str(row[0]).strip() if row[0] else ''
                empleado = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                nombres = str(row[2]).strip() if len(row) > 2 and row[2] else ''
                apellidos = str(row[3]).strip() if len(row) > 3 and row[3] else ''

                # Construir nombre completo combinando EMPLEADO, NOMBRES y APELLIDOS
                partes_nombre = [p for p in [empleado, nombres, apellidos] if p]
                nombre_completo = ' '.join(partes_nombre).strip()

                print(f"DEBUG: [FORMATO EXTENDIDO] RUT={rut}, EMPLEADO={empleado}, NOMBRES={nombres}, APELLIDOS={apellidos}")
                print(f"DEBUG: Nombre completo construido: '{nombre_completo}'")

                # Tipo de contrato desde columna específica o detectada
                if tipo_contrato_idx is not None and len(row) > tipo_contrato_idx:
                    tipo_contrato_raw = str(row[tipo_contrato_idx]).strip().lower()
                elif len(row) > 5:
                    tipo_contrato_raw = str(row[5]).strip().lower()
                else:
                    tipo_contrato_raw = 'indefinido'

            # Validaciones comunes
            if not rut:
                error = f"Fila {idx}: RUT vacío"
                print(f"DEBUG ERROR: {error}")
                errores.append(error)
                continue

            if not nombre_completo:
                error = f"Fila {idx}: Nombre vacío"
                print(f"DEBUG ERROR: {error}")
                errores.append(error)
                continue

            print(f"DEBUG: Tipo contrato raw: '{tipo_contrato_raw}'")

            # Mapear tipos de contrato
            if 'fijo' in tipo_contrato_raw or 'plazo' in tipo_contrato_raw:
                tipo_contrato = 'contratista'
            else:
                tipo_contrato = 'planta'

            print(f"DEBUG: Tipo contrato final: {tipo_contrato}")

            # Determinar tipo de caja
            if formato_simplificado:
                # En formato simplificado, tipo_caja está en la columna 3
                tipo_caja = tipo_caja_raw if tipo_caja_raw in ['estandar', 'especial', 'premium'] else 'estandar'
            else:
                # En formato extendido, usar valor por defecto
                tipo_caja = 'estandar'

            print(f"DEBUG: Tipo caja final: {tipo_caja}")

            # Determinar planta por fila (si existe columna), sino usar la planta por defecto
            planta_por_fila = planta
            try:
                # Obtener valor de planta desde la columna correspondiente
                raw_planta_val = ''
                if formato_simplificado:
                    # En formato simplificado, la planta_id está en columna 4
                    raw_planta_val = planta_id_raw
                elif planta_idx is not None and len(row) > planta_idx:
                    # En formato extendido, usar el índice detectado
                    raw_planta_val = str(row[planta_idx]).strip()

                if raw_planta_val:
                    print(f"DEBUG: Valor planta/sede del CSV: '{raw_planta_val}'")
                    # Intentar por ID numérico
                    try:
                        planta_id = int(raw_planta_val)
                        planta_por_fila = Planta.objects.get(id=planta_id)
                        print(f"DEBUG: Planta encontrada por ID {planta_id}: {planta_por_fila.nombre}")
                    except (ValueError, Planta.DoesNotExist):
                        # Si no es un ID numérico, mapear nombre/código a planta
                        raw_lower = raw_planta_val.lower()
                        planta_encontrada = False

                        # Mapeo de nombres comunes a códigos de planta
                        if 'santiago' in raw_lower or 'casablanca' in raw_lower or 'casa' in raw_lower or 'blanca' in raw_lower:
                            try:
                                planta_por_fila = Planta.objects.get(codigo='casablanca')
                                print(f"DEBUG: Mapeado '{raw_planta_val}' -> Casa Blanca")
                                planta_encontrada = True
                            except Planta.DoesNotExist:
                                pass
                        elif 'valparaiso' in raw_lower or 'valparaíso' in raw_lower:
                            if 'bic' in raw_lower:
                                try:
                                    planta_por_fila = Planta.objects.get(codigo='valparaiso_bic')
                                    print(f"DEBUG: Mapeado '{raw_planta_val}' -> Valparaíso BIC")
                                    planta_encontrada = True
                                except Planta.DoesNotExist:
                                    pass
                            else:
                                try:
                                    planta_por_fila = Planta.objects.get(codigo='valparaiso_bif')
                                    print(f"DEBUG: Mapeado '{raw_planta_val}' -> Valparaíso BIF")
                                    planta_encontrada = True
                                except Planta.DoesNotExist:
                                    pass
                        elif 'rancagua' in raw_lower:
                            # Rancagua: buscar si existe alguna planta asociada
                            print(f"DEBUG: '{raw_planta_val}' (Rancagua) - usando planta por defecto")

                        # Si no se encontró por mapeo, intentar búsqueda directa
                        if not planta_encontrada:
                            try:
                                planta_por_fila = Planta.objects.get(codigo=raw_lower)
                                print(f"DEBUG: Planta encontrada por código: {planta_por_fila.nombre}")
                            except Planta.DoesNotExist:
                                try:
                                    planta_por_fila = Planta.objects.get(nombre__iexact=raw_planta_val)
                                    print(f"DEBUG: Planta encontrada por nombre: {planta_por_fila.nombre}")
                                except Planta.DoesNotExist:
                                    print(f"DEBUG: No se pudo mapear '{raw_planta_val}' - usando planta por defecto")
            except Exception as ex:
                print(f"DEBUG: Error al determinar planta: {ex}")
                planta_por_fila = planta

            # Crear beneficiario
            print(f"DEBUG: Creando beneficiario '{nombre_completo}' ({rut}) en planta '{planta_por_fila}'...")
            try:
                beneficiario, created = Beneficiario.objects.get_or_create(
                    campana=campana,
                    rut=rut,
                    defaults={
                        'nombre': nombre_completo,
                        'tipo_contrato': tipo_contrato,
                        'tipo_caja': tipo_caja,
                        'planta': planta_por_fila
                    }
                )

                if created:
                    beneficiarios_creados += 1
                    print(f"DEBUG: ✓ Beneficiario creado exitosamente")
                else:
                    print(f"DEBUG: ⚠ Beneficiario ya existía")
            except Exception as e:
                error = f"Fila {idx}: Error al crear beneficiario - {str(e)}"
                print(f"DEBUG ERROR: {error}")
                errores.append(error)

        print(f"DEBUG: ==========================================")
        print(f"DEBUG: RESUMEN FINAL")
        print(f"DEBUG: Filas procesadas: {filas_procesadas}")
        print(f"DEBUG: Beneficiarios creados: {beneficiarios_creados}")
        print(f"DEBUG: Errores: {len(errores)}")
        print(f"DEBUG: ==========================================")

        if errores:
            print(f"DEBUG: DETALLE DE ERRORES:")
            for error in errores:
                print(f"  - {error}")

        # Validar que se haya creado al menos un beneficiario
        if beneficiarios_creados == 0:
            if errores:
                # Si hay errores, mostrar un resumen
                errores_muestra = errores[:5]  # Mostrar solo los primeros 5 errores
                mensaje_errores = "\n".join(f"• {e}" for e in errores_muestra)
                if len(errores) > 5:
                    mensaje_errores += f"\n... y {len(errores) - 5} errores más"
                raise Exception(f"No se pudo crear ningún beneficiario. Se encontraron {len(errores)} errores:\n{mensaje_errores}")
            else:
                raise Exception("No se encontraron datos válidos en el archivo. Asegúrese de que el archivo contenga al menos una fila con datos después del encabezado.")

    except Exception as e:
        print(f"DEBUG: !!!!! EXCEPCIÓN CAPTURADA !!!!!")
        print(f"DEBUG: {str(e)}")
        import traceback
        print(f"DEBUG: Traceback completo:")
        print(traceback.format_exc())
        raise

    return beneficiarios_creados


def _procesar_excel_nomina(archivo, campana, planta):
    """Procesa un archivo Excel"""
    try:
        wb = openpyxl.load_workbook(archivo)
    except Exception as e:
        raise Exception(f"Error al leer el archivo Excel. Asegúrese de que sea un archivo Excel válido (.xlsx o .xls): {str(e)}")

    ws = wb.active

    if ws is None:
        raise Exception("El archivo Excel no tiene hojas de cálculo.")

    beneficiarios_creados = 0
    errores = []
    filas_procesadas = 0

    # Leer encabezado (fila 1) para intentar detectar columna de planta
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)

    # Validar que exista encabezado
    if not header_row:
        raise Exception("El archivo Excel está vacío. No se encontró encabezado.")

    # Validar número de columnas (debe tener al menos 9)
    # Filtrar None values del encabezado para contar solo columnas con valor
    header_validas = [h for h in header_row if h is not None and str(h).strip()]
    if len(header_validas) < 9:
        raise Exception(f"El archivo tiene {len(header_validas)} columnas pero se requieren al menos 9 columnas: RUT, EMPLEADO, NOMBRES, APELLIDOS, CARGO, TIPO DE CONTRATO, PERIODO, SEDE, ESTADO")

    header_lower = [str(h).strip().lower() if h else '' for h in header_row]
    planta_idx = None
    for i, h in enumerate(header_lower):
        if any(k in h for k in ['planta', 'planta_id', 'sede', 'site', 'sucursal', 'centro']):
            planta_idx = i
            break
    print(f"DEBUG: Excel - índice de columna planta detectado: {planta_idx}")

    # Iterar desde la fila 2 (fila 1 es encabezado)
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        filas_procesadas += 1
        print(f"DEBUG Excel: Fila {idx}: {row}")

        # Saltar filas completamente vacías
        if not row or all(cell is None or str(cell).strip() == '' for cell in row):
            print(f"DEBUG Excel: Fila {idx} está vacía, saltando...")
            continue

        # Validar que tenga al menos los campos mínimos requeridos
        if len(row) < 4:
            error = f"Fila {idx}: Tiene {len(row)} columnas pero se requieren al menos 4 (RUT, EMPLEADO, NOMBRES, APELLIDOS)"
            print(f"DEBUG Excel ERROR: {error}")
            errores.append(error)
            continue

        # Validar que los campos requeridos no estén vacíos
        if not row[0] or str(row[0]).strip() == '':
            error = f"Fila {idx}: Campo RUT (columna 1) está vacío"
            print(f"DEBUG Excel ERROR: {error}")
            errores.append(error)
            continue

        try:
            # Formato Excel: RUT | EMPLEADO | NOMBRES | APELLIDOS | CARGO | TIPO DE CONTRATO | PERIODO | SEDE | ESTADO
            # 0: RUT
            # 1: EMPLEADO
            # 2: NOMBRES
            # 3: APELLIDOS
            # 4: CARGO
            # 5: TIPO DE CONTRATO
            # 6: PERIODO
            # 7: SEDE
            # 8: ESTADO

            rut = str(row[0]).strip() if row[0] else ''
            empleado = str(row[1]).strip() if len(row) > 1 and row[1] else ''
            nombres = str(row[2]).strip() if len(row) > 2 and row[2] else ''
            apellidos = str(row[3]).strip() if len(row) > 3 and row[3] else ''

            if not rut:
                error = f"Fila {idx}: Campo RUT está vacío"
                print(f"DEBUG Excel ERROR: {error}")
                errores.append(error)
                continue

            # Construir nombre completo combinando EMPLEADO, NOMBRES y APELLIDOS
            partes_nombre = [p for p in [empleado, nombres, apellidos] if p]
            nombre = ' '.join(partes_nombre).strip()

            if not nombre:
                error = f"Fila {idx}: Nombre completo vacío (EMPLEADO, NOMBRES y APELLIDOS están vacíos)"
                print(f"DEBUG Excel ERROR: {error}")
                errores.append(error)
                continue

            # Tipo de contrato (columna 5)
            tipo_contrato_raw = str(row[5]).strip().lower() if len(row) > 5 and row[5] else 'indefinido'

            # Mapear tipos de contrato
            if 'fijo' in tipo_contrato_raw or 'plazo' in tipo_contrato_raw:
                tipo_contrato = 'contratista'
            else:
                tipo_contrato = 'planta'

            # Tipo de caja por defecto
            tipo_caja = 'estandar'

            # Determinar planta por fila (si existe columna), sino usar la planta por defecto
            planta_por_fila = planta
            if planta_idx is not None and len(row) > planta_idx:
                raw_planta_val = str(row[planta_idx]).strip() if row[planta_idx] is not None else ''
                if raw_planta_val:
                    try:
                        planta_por_fila = Planta.objects.get(id=int(raw_planta_val))
                    except Exception:
                        try:
                            planta_por_fila = Planta.objects.get(codigo=raw_planta_val)
                        except Exception:
                            try:
                                planta_por_fila = Planta.objects.get(nombre__iexact=raw_planta_val)
                            except Exception:
                                planta_por_fila = planta

            # Crear beneficiario
            beneficiario, created = Beneficiario.objects.get_or_create(
                campana=campana,
                rut=rut,
                defaults={
                    'nombre': nombre,
                    'tipo_contrato': tipo_contrato,
                    'tipo_caja': tipo_caja,
                    'planta': planta_por_fila
                }
            )

            if created:
                beneficiarios_creados += 1
                print(f"DEBUG Excel: ✓ Beneficiario creado exitosamente")
            else:
                print(f"DEBUG Excel: ⚠ Beneficiario ya existía")
        except Exception as e:
            error = f"Fila {idx}: Error al crear beneficiario - {str(e)}"
            print(f"DEBUG Excel ERROR: {error}")
            errores.append(error)

    print(f"DEBUG Excel: ==========================================")
    print(f"DEBUG Excel: RESUMEN FINAL")
    print(f"DEBUG Excel: Filas procesadas: {filas_procesadas}")
    print(f"DEBUG Excel: Beneficiarios creados: {beneficiarios_creados}")
    print(f"DEBUG Excel: Errores: {len(errores)}")
    print(f"DEBUG Excel: ==========================================")

    if errores:
        print(f"DEBUG Excel: DETALLE DE ERRORES:")
        for error in errores:
            print(f"  - {error}")

    # Validar que se haya creado al menos un beneficiario
    if beneficiarios_creados == 0:
        if errores:
            # Si hay errores, mostrar un resumen
            errores_muestra = errores[:5]  # Mostrar solo los primeros 5 errores
            mensaje_errores = "\n".join(f"• {e}" for e in errores_muestra)
            if len(errores) > 5:
                mensaje_errores += f"\n... y {len(errores) - 5} errores más"
            raise Exception(f"No se pudo crear ningún beneficiario. Se encontraron {len(errores)} errores:\n{mensaje_errores}")
        else:
            raise Exception("No se encontraron datos válidos en el archivo. Asegúrese de que el archivo contenga al menos una fila con datos después del encabezado.")

    return beneficiarios_creados


def generar_excel_entregados(campana):
    """
    Genera un archivo Excel con la lista de beneficiarios que retiraron

    Returns:
        Workbook: Archivo Excel generado
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Entregados"

    # Encabezados
    ws.append(['Nombre', 'RUT', 'Tipo Contrato', 'Tipo Caja', 'Fecha Retiro', 'Hora Retiro', 'Confirmado Por'])

    # Datos
    beneficiarios = campana.beneficiarios.filter(retiro__isnull=False).select_related('retiro', 'retiro__confirmado_por')

    for b in beneficiarios:
        retiro = b.retiro
        ws.append([
            b.nombre,
            b.rut,
            b.get_tipo_contrato_display(),
            b.get_tipo_caja_display(),
            retiro.fecha_hora.strftime('%d/%m/%Y'),
            retiro.fecha_hora.strftime('%H:%M'),
            retiro.confirmado_por.get_full_name() if retiro.confirmado_por else 'N/A'
        ])

    return wb


def generar_excel_no_retirados(campana):
    """
    Genera un archivo Excel con la lista de beneficiarios que NO retiraron

    Returns:
        Workbook: Archivo Excel generado
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "No Retirados"

    # Encabezados
    ws.append(['Nombre', 'RUT', 'Tipo Contrato', 'Tipo Caja'])

    # Datos
    beneficiarios = campana.beneficiarios.filter(retiro__isnull=True)

    for b in beneficiarios:
        ws.append([
            b.nombre,
            b.rut,
            b.get_tipo_contrato_display(),
            b.get_tipo_caja_display()
        ])

    return wb


def validar_rut_chileno(rut):
    """
    Valida si un RUT chileno es válido.
    
    Acepta formatos:
    - 12.345.678-9
    - 12345678-9
    - 12345678
    - 123456789
    
    Args:
        rut (str): RUT a validar
        
    Returns:
        tuple: (es_valido: bool, mensaje: str)
    """
    if not rut:
        return False, "RUT no puede estar vacío"
    
    # Limpiar RUT: remover espacios, puntos, convertir a mayúsculas
    rut_limpio = str(rut).strip().replace('.', '').replace('-', '').upper()
    
    # Validar que solo contenga dígitos y opcionalmente K al final
    if not rut_limpio:
        return False, "RUT inválido: vacío después de limpiar"
    
    if not (rut_limpio[:-1].isdigit() and (rut_limpio[-1].isdigit() or rut_limpio[-1] == 'K')):
        return False, "RUT inválido: contiene caracteres no permitidos"
    
    # Validar longitud (7-8 dígitos + 1 dígito verificador = 8-9 caracteres total)
    if len(rut_limpio) < 8 or len(rut_limpio) > 9:
        return False, f"RUT inválido: longitud incorrecta (se esperan 8-9 caracteres, se recibieron {len(rut_limpio)})"
    
    # Separar números y dígito verificador
    numeros = rut_limpio[:-1]
    digito_proporcionado = rut_limpio[-1]
    
    # Calcular dígito verificador correcto
    suma = 0
    multiplicador = 2
    
    # Iterar desde la derecha hacia la izquierda
    for i in range(len(numeros) - 1, -1, -1):
        suma += int(numeros[i]) * multiplicador
        multiplicador += 1
        if multiplicador > 7:
            multiplicador = 2
    
    # Calcular el dígito verificador
    digito_calculado = 11 - (suma % 11)
    
    if digito_calculado == 11:
        digito_calculado = '0'
    elif digito_calculado == 10:
        digito_calculado = 'K'
    else:
        digito_calculado = str(digito_calculado)
    
    # Comparar
    if digito_proporcionado != digito_calculado:
        return False, f"RUT Inválido )"
    
    return True, "RUT válido"

